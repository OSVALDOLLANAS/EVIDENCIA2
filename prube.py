import sys
import csv
import openpyxl
from openpyxl import Workbook

biblioteca ={} 

try:
    with open("biblioteca.csv","r", newline="") as archivo:
        lector = csv.reader(archivo)
        next(lector)
    
        for identificador,titulo,autor,genero,año_publicacion,isbn,fecha_adquisicion in lector:
            biblioteca[int(identificador)] = (identificador,titulo,autor,genero,año_publicacion,isbn,fecha_adquisicion)
except FileNotFoundError:
    print("El archivo no se encontró, se procede a trabajar con un conjunto vacío")
except csv.Error as fallo_csv:
    print(f"Ocurrió un error al leer el archivo: {fallo_csv}")    
except Exception:
    Excepcion = sys.exc_info()
    print(f"Ocurrió un problema del tipo: {Excepcion[0]}")
    print(f"Mensaje del error: {Excepcion[1]}")
else:
    print(biblioteca)

while True:
    print("*"*40)
    print("****         MENU PRICIPAL         ****")
    opcion=int(input("[1]Registrar nuevo ejemplar\n[2]Consultas y reportes\n[3]Salir\nElija un opcion: "))
    print("*"*40)
    
    if opcion==1:
        print("*"*40)
        print("****  REGISTRO DE NUEVO EJEMPLEAR   ****")
        titulo=input("Ingrese el titulo: ")
        autor=input(f"Indique el autor de {titulo}: ")
        genero=input(f"Indique el genero de {titulo}: ")
        año_publicacion=input(f"Indique el año de publicación de {titulo}: ")
        isbn=input(f"Indique el ISBN de {titulo:}: ")
        fecha_adquisicion=input(f"Indique la fecha de adquisición de {titulo}: ")
        print("*"*40)
        identificador=max(biblioteca,default=0)+1
        biblioteca[identificador]=[identificador,titulo,autor,genero,año_publicacion,isbn,fecha_adquisicion]
        archivo=open("biblioteca.csv","w", newline="")
        guardado= csv.writer(archivo)
        guardado.writerows([(int(identificador), datos[1], datos[2],datos[3],datos[4],datos[5],datos[6]) for identificador, datos in biblioteca.items()])
        archivo.close()
        continue

    if opcion==2:
        while True:
            print("*"*40)
            print("****      CONSULTAS Y REPORTES      ****")
            opcion_2=int(input("[1]Consulta de titulo\n[2]Reportes\n[3]Volver al menu de consultas y reportes\nElija una opcion: "))
            print("*"*40)

            if opcion_2==1:
                while True:
                    print("*"*40)
                    opcion2_1=int(input("[1]Por titulo\n[2]Por ISBN\n[3]Volver al menu de consultas y reportes\nElija una opcion: "))

                    if opcion2_1==1:
                        print("*"*40)
                        print(f"ID\tTITULO")
                        print("*"*40)
                        for identificador,titulo,autor,genero,año_publicacion,isbn,fecha_adquisicion in biblioteca.values():
                            print(f"{identificador:3} | {titulo}")
                        else:
                            print("*"*40)
                        print("")
                        titulo_consulta=input("Elija una opcion para obtener todos los datos de el titulo: ")
                        print("")
                        print("*"*130)
                
                    if opcion2_1==2:
                        print("*"*40)
                        print(f"ID\tISBN")
                        print("*"*40)
                        for identificador,titulo,autor,genero,año_publicacion,isbn,fecha_adquisicion in biblioteca.values():
                            print(f"{identificador:3} | {isbn}")
                        else:
                            print("*"*40)
                            print("")
                        isbn_consulta=input("Elija una opcion para obtener todos los datos de el isbn: ")
                        print("")
                        print("*"*130)

                    if opcion2_1==3:
                        break

            if opcion_2==2:
                while True:
                    print("*"*40)
                    print("****            REPORTES            ****")
                    print("*"*40)
                    opcion2_2=int(input("[1]Catalago completo\n[2]Reporte por autor\n[3]Reporte por género\n[4]Por año de publicación\n[5]Volver al menú de reportes\nElija una opcion: "))
                    print("")

                    if opcion2_2==1:
                        print("*"*130)
                        print(f"ID\tTITULO\t\t\tAUTOR\t\tGENERO\t\tAÑO DE PUBLICACION\tISBN\t\tFECHA DE ADQUISICIÓN")
                        print("*"*130)
                        for identificador,titulo,autor,genero,año_publicacion,isbn,fecha_adquisicion in biblioteca.values():
                            print(f"{identificador:3}|{titulo:20}|{autor:20}|{genero:15}|{año_publicacion:20}|{isbn:20}|{fecha_adquisicion}")
                        else:
                            print("*"*130)
                        print("")

                        while True:
                            opcion2_2_1=int(input("[1]Exportar reporte en formato CSV\n[2]Exportar reporte en formato MsExcel\n[3]No exportar nada"))

                            if opcion2_2_1==1:
                                with open("biblioteca.csv","w",newline="") as archivo:
                                    guardado=csv.writer(archivo)
                                    guardado.writerow(biblioteca.keys())
                                    guardado.writerow(biblioteca.values())
                                print("*"*40)
                                print("El reporte se guardo en el archivo biblioteca.csv")
                                print("*"*40)
                            if opcion2_2_1==2:
                                libro = openpyxl.Workbook()
                                hoja = libro.active 
                                hoja["B1"].value = "BIBLIOTECA"
                                for i, key in enumerate(biblioteca.keys(),start=1):
                                    hoja.cell(row=1, column=i, value=key)
                                libro.save("biblioteca.xlsx")
                                print("*"*40)
                                print("El reporte se guardo en el archivo biblioteca.xlsx")
                                print("*"*40)
                            if opcion2_2_1==3:
                                break   
                        
                    if opcion2_2==2:
                        print("*"*40)
                        print(f"ID\tAUTOR")
                        print("*"*40)
                        for identificador,titulo,autor,genero,año_publicacion,isbn,fecha_adquisicion in biblioteca.values():
                            print(f"{identificador:3}|{autor}")
                        else:
                            print("*"*40)
                        print("")
                        autor_opcion=input(f"Elija un autor para mostrar todas los ejemplares existentes: ")
                        print("")
                        print("*"*130)
                        print(f"ID\tTITULO\t\t\tAUTOR\t\tGENERO\t\tAÑO DE PUBLICACION\tISBN\t\tFECHA DE ADQUISICIÓN")
                        print("*"*130)
                        for identificador,titulo,autor,genero,año_publicacion,isbn,fecha_adquisicion in biblioteca.values():
                            if autor==autor_opcion:
                                print(f"{identificador:3}|{titulo:20}|{autor:20}|{genero:15}|{año_publicacion:20}|{isbn:20}|{fecha_adquisicion}")
                        else:
                            print("*"*130)
                        print("")

                        while True:
                            opcion2_2_1=int(input("[1]Exportar reporte en formato CSV\n[2]Exportar reporte en formato MsExcel\n[3]No exportar nada"))
                            if opcion2_2_1==1:
                                with open("biblioteca.csv","w",newline="") as archivo:
                                    guardado=csv.writer(archivo)
                                    guardado.writerow(biblioteca.keys())
                                    guardado.writerow(biblioteca.values())
                                print("*"*40)
                                print("El reporte se guardo en el archivo biblioteca.csv")
                                print("*"*40)
                            if opcion2_2_1==2:
                                libro = openpyxl.Workbook()
                                hoja = libro.active 
                                hoja["B1"].value = "BIBLIOTECA"
                                for i, key in enumerate(biblioteca.keys(),start=1):
                                    hoja.cell(row=1, column=i, value=key)
                                libro.save("biblioteca.xlsx")
                                print("*"*40)
                                print("El reporte se guardo en el archivo biblioteca.xlsx")
                                print("*"*40)
                            if opcion2_2_1==3:
                                break   


                    if opcion2_2==3:
                        print("*"*40)
                        print(f"ID\tGENERO")
                        print("*"*40)
                        for identificador,titulo,autor,genero,año_publicacion,isbn,fecha_adquisicion in biblioteca.values():
                            print(f"{identificador:3}|{genero}")
                        else:
                            print("*"*40)
                        print("")
                        genero_opcion=input(f"Elija un genero para mostrar todas los ejemplares existentes: ")
                        print("")
                        print("*"*130)
                        print(f"ID\tTITULO\t\t\tAUTOR\t\tGENERO\t\tAÑO DE PUBLICACION\tISBN\t\tFECHA DE ADQUISICIÓN")
                        print("*"*130)
                        for identificador,titulo,autor,genero,año_publicacion,isbn,fecha_adquisicion in biblioteca.values():
                            if genero==genero_opcion:
                                print(f"{identificador:3}|{titulo:20}|{autor:20}|{genero:15}|{año_publicacion:20}|{isbn:20}|{fecha_adquisicion}")
                        else:
                            print("*"*130)
                        print("")

                        while True:
                            opcion2_2_1=int(input("[1]Exportar reporte en formato CSV\n[2]Exportar reporte en formato MsExcel\n[3]No exportar nada"))
                            if opcion2_2_1==1:
                                with open("biblioteca.csv","w",newline="") as archivo:
                                    guardado=csv.writer(archivo)
                                    guardado.writerow(biblioteca.keys())
                                    guardado.writerow(biblioteca.values())
                                print("*"*40)
                                print("El reporte se guardo en el archivo biblioteca.csv")
                                print("*"*40)
                            if opcion2_2_1==2:
                                libro = openpyxl.Workbook()
                                hoja = libro.active 
                                hoja["B1"].value = "BIBLIOTECA"
                                for i, key in enumerate(biblioteca.keys(),start=1):
                                    hoja.cell(row=1, column=i, value=key)
                                libro.save("biblioteca.xlsx")
                                print("*"*40)
                                print("El reporte se guardo en el archivo biblioteca.xlsx")
                                print("*"*40)
                            if opcion2_2_1==3:
                                break   

                    if opcion2_2==4:
                        print("*"*40)
                        print(f"ID\tAÑO ESPECIFICO")
                        print("*"*40)
                        for identificador,titulo,autor,genero,año_publicacion,isbn,fecha_adquisicion in biblioteca.values():
                            print(f"{identificador:3}|{año_publicacion}")
                        else:
                            print("*"*40)
                        print("")
                        año_publicacion_opcion=input(f"Elija un año para mostrar todas los ejemplares existentes: ")
                        print("")
                        print("*"*130)
                        print(f"ID\tTITULO\t\t\tAUTOR\t\tGENERO\t\tAÑO DE PUBLICACION\tISBN\t\tFECHA DE ADQUISICIÓN")
                        print("*"*130)
                        for identificador,titulo,autor,genero,año_publicacion,isbn,fecha_adquisicion in biblioteca.values():
                            if año_publicacion==año_publicacion_opcion:
                                print(f"{identificador:3}|{titulo:20}|{autor:20}|{genero:15}|{año_publicacion:20}|{isbn:20}|{fecha_adquisicion}")
                        else:
                            print("*"*130)
                        print("")

                        while True:
                            opcion2_2_1=int(input("[1]Exportar reporte en formato CSV\n[2]Exportar reporte en formato MsExcel\n[3]No exportar nada"))
                            if opcion2_2_1==1:
                                with open("biblioteca.csv","w",newline="") as archivo:
                                    guardado=csv.writer(archivo)
                                    guardado.writerow(biblioteca.keys())
                                    guardado.writerow(biblioteca.values())
                                print("*"*40)
                                print("El reporte se guardo en el archivo biblioteca.csv")
                                print("*"*40)
                            if opcion2_2_1==2:
                                libro = openpyxl.Workbook()
                                hoja = libro.active 
                                hoja["B1"].value = "BIBLIOTECA"
                                for i, key in enumerate(biblioteca.keys(),start=1):
                                    hoja.cell(row=1, column=i, value=key)
                                libro.save("biblioteca.xlsx")
                                print("*"*40)
                                print("El reporte se guardo en el archivo biblioteca.xlsx")
                                print("*"*40)
                            if opcion2_2_1==3:
                                break   
                    if opcion2_2==5:
                        break 
            if opcion_2==3:
                break  
    if opcion==3:
        print("Archivo CSV tiene como nombre 'biblioteca.csv")
        print(archivo)
        archivo.close()
        break